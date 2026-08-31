#!/usr/bin/env python3
"""Generate ready-to-run cURL commands and a QA matrix from every available source.

Reads whatever is present rather than requiring one canonical spec:

    api-docs/API_File.json          the inventory: method, base URL, path, body, headers
    collections/**/*.json           Postman collections
    <external>/**/*.yml             an OpenCollection/Bruno-style export, if cloned
    attendance-management/*.xlsx    the spec workbook, where rows are populated

Endpoints are keyed on method + normalised path so the same endpoint described in
two sources becomes one entry with the union of what they know, rather than two
near-duplicates. The source of each field is recorded, because a payload that
came from a working export and one that came from a template are not equally
trustworthy and the reader should be able to tell them apart.

No credential is ever written out. Authorization becomes "Bearer $TOKEN" and the
generated file carries a line telling the reader how to populate it, so the
output can be committed and shared without a scrub step -- a generator that
sometimes emits a live token is a generator nobody can safely run.
"""
from __future__ import annotations

import glob
import io
import json
import os
import re
import sys
from collections import OrderedDict, defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXTERNAL = os.environ.get(
    "EXTERNAL_COLLECTION_DIR",
    os.path.join(os.path.dirname(ROOT), "attendance_apis_ro"),
)

#: Header names whose value is a credential. Emitted as a placeholder.
SECRET_HEADERS = re.compile(
    r"^(authorization|cookie|x-api-key|x-auth-token|proxy-authorization)$", re.I
)

#: Anything of this shape is a variable the runner substitutes, not a secret,
#: so it survives into the output as-is.
TEMPLATE = re.compile(r"\{\{[^}]+\}\}")

#: Body fields that are credentials in their own right. Stripping only the
#: Authorization header leaves the login payload carrying an employee code and
#: password in clear, and this file is meant to be committed and shared.
#: Replaced with shell variables so the command still runs as written.
SECRET_FIELDS = {
    "empcode": "$EMP_CODE",
    "password": "$PASSWORD",
    "emppassword": "$PASSWORD",
    "secret": "$SECRET",
    "clientsecret": "$SECRET",
    "apikey": "$API_KEY",
    "accesstoken": "$TOKEN",
}


def redact_body(text):
    """Swap credential values in a JSON payload for shell variables.

    Non-JSON bodies are returned untouched: a blind regex over arbitrary text
    is more likely to corrupt a payload than to find a secret in it.
    """
    if not text:
        return text
    try:
        doc = json.loads(text)
    except Exception:
        return text

    def walk(o):
        if isinstance(o, dict):
            out = {}
            for k, v in o.items():
                flat = str(k).lower().replace("_", "").replace("-", "")
                out[k] = SECRET_FIELDS[flat] if flat in SECRET_FIELDS and isinstance(v, str) \
                    else walk(v)
            return out
        if isinstance(o, list):
            return [walk(v) for v in o]
        return o

    return json.dumps(walk(doc), separators=(",", ":"))


def norm_path(p: str) -> str:
    """Collapse concrete ids so /policy/2 and /policy/{id} are one endpoint."""
    p = re.sub(r"\{[^}]+\}", "{id}", p or "")
    p = re.sub(r"/-?\d+(?=/|$)", "/{id}", p)
    return p.rstrip("/").lower() or "/"


class Endpoint:
    __slots__ = ("method", "path", "base", "name", "module", "headers",
                 "body", "sources", "ids", "expected")

    def __init__(self, method, path, base=None, name="", module=""):
        self.method = (method or "GET").upper()
        self.path = path
        self.base = base
        self.name = name
        self.module = module
        self.headers = OrderedDict()
        self.body = None
        self.sources = []
        self.ids = set()
        self.expected = set()

    @property
    def key(self):
        return (self.method, norm_path(self.path))

    def merge(self, other):
        """Fill only what is missing. First source to supply a field wins, and
        the sources list records who contributed, so a later template cannot
        quietly overwrite a payload taken from a real request."""
        self.base = self.base or other.base
        self.name = self.name or other.name
        self.module = self.module or other.module
        self.body = self.body if self.body else other.body
        for k, v in other.headers.items():
            self.headers.setdefault(k, v)
        self.ids |= other.ids
        self.expected |= other.expected
        for s in other.sources:
            if s not in self.sources:
                self.sources.append(s)


def safe_headers(raw: dict) -> OrderedDict:
    out = OrderedDict()
    for k, v in (raw or {}).items():
        if not k:
            continue
        if SECRET_HEADERS.match(k.strip()):
            out[k.strip()] = "Bearer $TOKEN" if k.strip().lower() == "authorization" else "$SECRET"
        else:
            out[k.strip()] = "" if v is None else str(v).strip()
    return out


# --------------------------------------------------------------- sources ----
def from_inventory(path):
    try:
        rows = json.load(io.open(path, encoding="utf-8"))
    except Exception:
        return []
    rows = rows if isinstance(rows, list) else rows.get("apis", [])
    out = []
    for r in rows:
        p = str(r.get("Endpoint / Path") or "").strip()
        if not p:
            continue
        e = Endpoint(r.get("HTTP Method"), p,
                     str(r.get("Base URL") or "").strip() or None,
                     str(r.get("API Name") or r.get("Name") or "").strip(),
                     str(r.get("Module") or "").strip())
        params = str(r.get("Request Parameters") or "")
        hdrs = {}
        m = re.search(r"headers:\s*([^|]+)", params, re.I)
        if m:
            for pair in m.group(1).split(";"):
                if "=" in pair:
                    k, v = pair.split("=", 1)
                    hdrs[k.strip()] = v.strip()
        e.headers = safe_headers(hdrs)
        body = str(r.get("Request Body") or "").strip()
        if body:
            e.body = body
        for s in re.findall(r"\b([1-5]\d\d)\b", str(r.get("Expected Status") or "")):
            e.expected.add(s)
        e.sources.append("inventory")
        out.append(e)
    return out


def from_postman(pattern):
    out = []
    for f in glob.glob(pattern, recursive=True):
        try:
            doc = json.load(io.open(f, encoding="utf-8"))
        except Exception:
            continue
        cname = (doc.get("info") or {}).get("name") or os.path.basename(f)

        def walk(items):
            for it in items:
                if isinstance(it.get("item"), list):
                    walk(it["item"])
                    continue
                req = it.get("request")
                if not req:
                    continue
                url = req.get("url")
                raw = url if isinstance(url, str) else (url or {}).get("raw") or ""
                if not raw:
                    continue
                base, _, tail = raw.partition("://")
                if tail:
                    host, _, rest = tail.partition("/")
                    origin, p = f"{base}://{host}", "/" + rest
                else:
                    origin, p = None, raw
                e = Endpoint(req.get("method"), p.split("?")[0], origin,
                             it.get("name") or "", cname)
                e.headers = safe_headers(
                    {h.get("key"): h.get("value") for h in (req.get("header") or [])}
                )
                b = (req.get("body") or {}).get("raw")
                if b:
                    e.body = b.strip()
                e.sources.append("postman")
                out.append(e)

        walk(doc.get("item") or [])
    return out


def from_yaml_collection(root):
    """An OpenCollection/Bruno-style export: one request per .yml."""
    out = []
    if not os.path.isdir(root):
        return out
    for f in glob.glob(os.path.join(root, "**", "*.yml"), recursive=True):
        if os.sep + "reports" + os.sep in f:
            continue
        try:
            t = io.open(f, encoding="utf-8-sig", errors="replace").read()
        except Exception:
            continue
        mu = re.search(r"^\s*url:\s*(\S+)", t, re.M)
        mm = re.search(r"^\s*method:\s*(\w+)", t, re.M)
        if not mu:
            continue
        url = mu.group(1)
        origin = None
        om = re.match(r"(https?://[^/]+)(/.*)?$", url)
        p = url
        if om:
            origin, p = om.group(1), om.group(2) or "/"
        p = p.split("?")[0]
        e = Endpoint(mm.group(1) if mm else "GET", p, origin,
                     (re.search(r"name:\s*(.+)", t) or [None, ""])[1].strip(),
                     os.path.basename(os.path.dirname(f)))
        hdrs = {}
        for hm in re.finditer(r"-\s*name:\s*(\S+)\s*\n\s*value:\s*(.*)", t):
            hdrs[hm.group(1)] = hm.group(2).strip()
        e.headers = safe_headers(hdrs)
        bm = re.search(r"^\s*body:\s*\n((?:\s{2,}.*\n)+)", t, re.M)
        if bm:
            body = re.sub(r"^\s{0,6}", "", bm.group(1), flags=re.M).strip()
            if body.startswith("{"):
                e.body = body
        for i in re.findall(r"/(-?\d+)(?=/|$)", p):
            e.ids.add(i)
        e.sources.append("yaml-collection")
        out.append(e)
    return out


def from_workbook(path):
    try:
        import openpyxl
    except ImportError:
        return []
    if not os.path.exists(path):
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return []

    def rows(sheet):
        if sheet not in wb.sheetnames:
            return []
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        if not data:
            return []
        hdr = [str(c).strip() if c is not None else "" for c in data[0]]
        out = []
        for r in data[1:]:
            if not any(c is not None and str(c).strip() for c in r):
                continue          # max_row counts formatting; skip empty rows
            out.append({hdr[i]: r[i] for i in range(min(len(hdr), len(r)))})
        return out

    payloads = defaultdict(dict)
    for r in rows("Sample_Payloads"):
        payloads[str(r.get("API ID") or "").strip()][
            str(r.get("Payload Type") or "").strip().lower()
        ] = r.get("Sample JSON")

    out = []
    for r in rows("API_Overview"):
        p = str(r.get("Endpoint Path") or "").strip()
        if not p:
            continue
        e = Endpoint(r.get("HTTP Method"), p,
                     str(r.get("Base URL") or "").strip() or None,
                     str(r.get("API / Feature Name") or "").strip(),
                     str(r.get("Module") or "").strip())
        if str(r.get("Auth Type") or "").lower().find("bearer") >= 0:
            e.headers["Authorization"] = "Bearer $TOKEN"
        e.headers.setdefault("Content-Type", "application/json")
        body = payloads.get(str(r.get("API ID") or "").strip(), {}).get("request body")
        if body:
            e.body = str(body).strip()
        e.sources.append("workbook")
        out.append(e)
    return out


# ----------------------------------------------------------------- build ----
def collect():
    found = []
    found += from_inventory(os.path.join(ROOT, "api-docs", "API_File.json"))
    found += from_postman(os.path.join(ROOT, "collections", "**", "*.json"))
    found += from_yaml_collection(EXTERNAL)
    found += from_workbook(os.path.join(
        ROOT, "attendance-management", "Attendance_Management_API_Spec.xlsx"))

    merged = OrderedDict()
    for e in found:
        if e.key in merged:
            merged[e.key].merge(e)
        else:
            merged[e.key] = e
    return list(merged.values()), found


def curl(e, scenario, override_headers=None, body=None, path=None):
    p = path if path is not None else e.path
    url = (e.base or "{{baseUrl}}") + p
    parts = [f"curl -sS -X {e.method} '{url}'"]
    hdrs = OrderedDict(e.headers if override_headers is None else override_headers)
    if e.method in ("POST", "PUT", "PATCH") and body and "Content-Type" not in hdrs:
        hdrs["Content-Type"] = "application/json"
    for k, v in hdrs.items():
        parts.append(f"  -H '{k}: {v}'")
    if body:
        body = redact_body(body)
        compact = json.dumps(json.loads(body), separators=(",", ":")) \
            if _is_json(body) else body.replace("\n", " ")
        parts.append("  -d '" + compact.replace("'", "'\\''") + "'")
    return " \\\n".join(parts)


def _is_json(t):
    try:
        json.loads(t)
        return True
    except Exception:
        return False


def scenarios(e):
    """Positive plus the negative cases that apply to every authenticated API.

    Deliberately not a per-endpoint invention: these are the global cases the
    contract tier already runs, written out so they can be reproduced by hand.
    """
    out = [("Positive — documented request", dict(headers=None, body=e.body,
                                                  status="2xx"))]
    if any(k.lower() == "authorization" for k in e.headers):
        no_auth = OrderedDict((k, v) for k, v in e.headers.items()
                              if k.lower() != "authorization")
        out.append(("Negative — no credential", dict(headers=no_auth, body=e.body,
                                                     status="401 or 403")))
        bad = OrderedDict(e.headers)
        bad["Authorization"] = "Bearer not-a-real-token"
        out.append(("Negative — malformed token", dict(headers=bad, body=e.body,
                                                       status="401")))
    if e.method in ("POST", "PUT", "PATCH"):
        out.append(("Negative — empty body", dict(headers=None, body="{}",
                                                  status="400 or 422")))
        if e.body:
            out.append(("Negative — wrong content type",
                        dict(headers=OrderedDict(list(e.headers.items()) +
                                                 [("Content-Type", "text/plain")]),
                             body=e.body, status="415")))
    if "{id}" in norm_path(e.path):
        out.append(("Negative — non-existent id",
                    dict(headers=None, body=e.body, status="404",
                         path=re.sub(r"/-?\d+(?=/|$)", "/999999",
                                     re.sub(r"\{[^}]+\}", "999999", e.path)))))
    return out


def main():
    endpoints, raw = collect()
    endpoints.sort(key=lambda e: (e.module or "", e.path, e.method))

    out_dir = os.path.join(ROOT, "docs", "curl")
    os.makedirs(out_dir, exist_ok=True)

    total_cases = 0
    lines = [
        "# cURL command matrix",
        "",
        "Generated by `scripts/generate_curl_matrix.py` from every source present:",
        "the inventory, the Postman collections, an OpenCollection export if one is",
        "cloned, and the spec workbook.",
        "",
        "**No credential appears below.** `Authorization` is emitted as `Bearer $TOKEN`,",
        "and credential fields inside request bodies -- an employee code, a password --",
        "are replaced the same way. Populate them before running:",
        "",
        "```sh",
        "TOKEN=$(curl -sS -X POST 'https://uat-mcdp-be.omfysgroup.com/auth/token' \\",
        "  -H 'Content-Type: application/json' \\",
        "  -d '{\"empCode\":\"<EMP_CODE>\",\"password\":\"<PASSWORD>\"}' | jq -r '.token')",
        "```",
        "",
    ]

    matrix = [("Module", "Method", "Endpoint", "Scenario", "Expected", "Sources")]
    by_module = defaultdict(list)
    for e in endpoints:
        by_module[e.module or "(unattributed)"].append(e)

    for module, eps in sorted(by_module.items()):
        lines += [f"## {module}", ""]
        for e in eps:
            lines += [f"### {e.method} {e.path}", ""]
            if e.name:
                lines.append(f"*{e.name}*")
            lines.append(f"<sub>source: {', '.join(e.sources)}"
                         f"{' · ids seen: ' + ', '.join(sorted(e.ids)) if e.ids else ''}</sub>")
            lines.append("")
            for title, cfg in scenarios(e):
                total_cases += 1
                lines += [f"**{title}** — expect `{cfg['status']}`", "", "```sh",
                          curl(e, title, cfg.get("headers"), cfg.get("body"),
                               cfg.get("path")),
                          "```", ""]
                matrix.append((module, e.method, e.path, title, cfg["status"],
                               "/".join(e.sources)))

    io.open(os.path.join(out_dir, "CURL_MATRIX.md"), "w",
            encoding="utf-8", newline="\n").write("\n".join(lines))

    with io.open(os.path.join(out_dir, "qa_matrix.csv"), "w",
                 encoding="utf-8", newline="\n") as fh:
        for row in matrix:
            fh.write(",".join('"' + str(c).replace('"', '""') + '"' for c in row) + "\n")

    print(f"  merged endpoints : {len(endpoints)}  (from {len(raw)} source entries)")
    print(f"  cURL commands    : {total_cases}")
    print(f"  written          : docs/curl/CURL_MATRIX.md, docs/curl/qa_matrix.csv")
    contrib = defaultdict(int)
    for e in endpoints:
        for s in e.sources:
            contrib[s] += 1
    print("  contributing sources:")
    for k, v in sorted(contrib.items(), key=lambda kv: -kv[1]):
        print(f"    {k:18} {v:3} endpoint(s)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
