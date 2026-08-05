import openpyxl
import os

wb = openpyxl.Workbook()

# Sheet 1: API_Overview
ws1 = wb.active
ws1.title = "API_Overview"
ws1.append([
    "API ID", "API / Feature Name", "Module", "Purpose", "Owner / Dev Contact",
    "HTTP Method", "Base URL", "Endpoint Path", "Auth Type", "Idempotent (Y/N)",
    "Environment(s)", "API Version", "Last Updated", "Postman Collection Link"
])

overview_rows = [
    ["ATT-001", "Mark Check-In", "HCM - Attendance", "Record daily employee check-in timestamp and location", "QA Team", "POST", "{{baseUrl}}", "/user/attendance/checkin", "Bearer Token", "No", "Dev, QA, UAT", "v1", "2026-08-05", "attendance-management/collections/Attendance_API.json"],
    ["ATT-002", "Mark Check-Out", "HCM - Attendance", "Record daily employee check-out timestamp and location", "QA Team", "POST", "{{baseUrl}}", "/user/attendance/checkout", "Bearer Token", "No", "Dev, QA, UAT", "v1", "2026-08-05", "attendance-management/collections/Attendance_API.json"],
    ["ATT-003", "Get Attendance History", "HCM - Attendance", "Retrieve historical attendance logs for employee", "QA Team", "GET", "{{baseUrl}}", "/user/attendance/history", "Bearer Token", "Yes", "Dev, QA, UAT", "v1", "2026-08-05", "attendance-management/collections/Attendance_API.json"],
    ["ATT-004", "Get Attendance Summary", "HCM - Attendance", "Get monthly summary metrics (present, absent, late counts)", "QA Team", "GET", "{{baseUrl}}", "/user/attendance/summary", "Bearer Token", "Yes", "Dev, QA, UAT", "v1", "2026-08-05", "attendance-management/collections/Attendance_API.json"],
    ["ATT-005", "Apply Regularization", "HCM - Attendance", "Submit attendance regularization request for missed check-in/out", "QA Team", "POST", "{{baseUrl}}", "/user/attendance/regularize", "Bearer Token", "No", "Dev, QA, UAT", "v1", "2026-08-05", "attendance-management/collections/Attendance_API.json"]
]
for r in overview_rows:
    ws1.append(r)

# Sheet 2: Sample_Payloads
ws2 = wb.create_sheet(title="Sample_Payloads")
ws2.append(["API ID", "Payload Type", "Sample JSON"])

payload_rows = [
    ["ATT-001", "Request Body", '{\n  "empCode": "OMI-0076",\n  "checkInTime": "09:00:00",\n  "location": "Office HQ"\n}'],
    ["ATT-001", "Success Response (200)", '{\n  "status": "SUCCESS",\n  "message": "Check-in recorded successfully",\n  "checkInTime": "09:00:00"\n}'],
    ["ATT-001", "Error Response (400)", '{\n  "errorCode": "DUPLICATE_CHECKIN",\n  "message": "Employee already checked in for today"\n}'],
    ["ATT-002", "Request Body", '{\n  "empCode": "OMI-0076",\n  "checkOutTime": "18:00:00",\n  "location": "Office HQ"\n}'],
    ["ATT-002", "Success Response (200)", '{\n  "status": "SUCCESS",\n  "message": "Check-out recorded successfully",\n  "totalHours": 9.0\n}'],
    ["ATT-003", "Success Response (200)", '{\n  "records": [\n    {\n      "date": "2026-08-01",\n      "checkIn": "09:02:11",\n      "checkOut": "18:05:40",\n      "status": "PRESENT"\n    }\n  ]\n}'],
    ["ATT-004", "Success Response (200)", '{\n  "empCode": "OMI-0076",\n  "totalWorkingDays": 22,\n  "presentCount": 20,\n  "absentCount": 1,\n  "leaveCount": 1,\n  "lateCount": 2\n}'],
    ["ATT-005", "Request Body", '{\n  "empCode": "OMI-0076",\n  "date": "2026-08-04",\n  "requestedCheckIn": "09:00:00",\n  "requestedCheckOut": "18:00:00",\n  "reason": "Biometric device failure"\n}'],
    ["ATT-005", "Success Response (200)", '{\n  "regularizationId": "REG-2026-0892",\n  "status": "PENDING_APPROVAL"\n}']
]
for r in payload_rows:
    ws2.append(r)

# Sheet 3: Rules_Dependencies_EdgeCases
ws3 = wb.create_sheet(title="Rules_Dependencies_EdgeCases")
ws3.append(["API ID", "Category", "Description"])

rules_rows = [
    ["ATT-001", "Business Rule", "Multiple check-ins on the same day must be rejected with 400 Bad Request / DUPLICATE_CHECKIN."],
    ["ATT-002", "Business Rule", "Check-out prior to a valid check-in must return 400 Bad Request / NO_CHECKIN_FOUND."],
    ["ATT-003", "Business Rule", "Attendance history defaults to current month if no date range parameters provided."],
    ["ATT-004", "Business Rule", "Summary statistics calculate present, absent, leave, and late marks according to official shift timings."],
    ["ATT-005", "Business Rule", "Maximum of 3 attendance regularization requests allowed per month per employee."],
    ["ATT-ALL", "Dependency", "Requires valid Bearer JWT authToken obtained from POST /auth/token."],
    ["ATT-ALL", "Dependency", "Employee must exist in system with ACTIVE status."],
    ["ATT-001", "Edge Case", "Cross-midnight shift check-in/out handling."],
    ["ATT-003", "Edge Case", "Inverted date range parameter query (startDate > endDate) must return empty records or 400."],
    ["ATT-ALL", "Non-Functional", "Expected response latency < 500ms under normal load; TLS certificate pinning enforced."]
]
for r in rules_rows:
    ws3.append(r)

target_path = "E:/API_Automation-main/attendance-management/Attendance_Management_API_Spec.xlsx"
wb.save(target_path)
print(f"SUCCESS: Created {target_path}")
