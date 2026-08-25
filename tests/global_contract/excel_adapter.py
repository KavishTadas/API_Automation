"""Excel upload -> API definition.

Parses ``api-docs/API_Documentation_Template.xlsx`` (or any workbook with the
same three sheets) into :class:`~tests.global_contract.metadata_resolver.ApiDefinition`
objects keyed by API ID, and projects each one into an ``API_File.json``-shaped
request row so ``perform_api_request()`` works unchanged.

Nothing is written back. ``api-docs/API_File.json`` is a generated artifact and
stays untouched — an uploaded API lives only for the duration of its run (DR-2).

Sheet contract
--------------
``API_Overview`` (15 columns) carries one row per API. ``Sample_Payloads``
(4 columns) carries four row types per API ID, and ``Rules_Dependencies_EdgeCases``
(3 columns) carries free prose that is deliberately **not** parsed — in
particular no SLA is mined out of "Expected response time < 500ms", which reads
like structured data but is a human note.

Only ``attendance-management/API_Documentation_Template.xlsx`` has the older
14-column shape; that path is quarantined and must not be read from.
"""

from __future__ import annotations

from dataclasses import replace
from pathlib import Path
from typing import Any

from tests.global_contract.metadata_resolver import (
    ROOT_DIR,
    ApiDefinition,
    ApiRule,
    PayloadType,
    SamplePayload,
    _coerce_status,
    _parse_json_or_none,
    definition_to_inventory_row,
    error_trigger_rows,
)


__all__ = [
    "ExcelAdapterError",
    "TEMPLATE_PATH",
    "definition_to_inventory_row",
    "error_trigger_rows",
    "load_excel_definitions",
    "parse_workbook",
]


#: The canonical template location. The 14-column copy under
#: ``attendance-management/`` is quarantined and is never read.
TEMPLATE_PATH = ROOT_DIR / "api-docs" / "API_Documentation_Template.xlsx"

OVERVIEW_SHEET = "API_Overview"
PAYLOADS_SHEET = "Sample_Payloads"
RULES_SHEET = "Rules_Dependencies_EdgeCases"

#: `API_Overview` header -> ApiDefinition field.
_OVERVIEW_COLUMNS = {
    "API ID": "api_id",
    "API / Feature Name": "name",
    "Module": "module",
    "Purpose": "purpose",
    "Owner / Dev Contact": "owner",
    "HTTP Method": "method",
    "Base URL": "base_url",
    "Endpoint Path": "path",
    "Auth Type": "auth_type",
    "Idempotent (Y/N)": "idempotent",
    "Environment(s)": "environments",
    "API Version": "api_version",
    "Last Updated": "last_updated",
    "cURL": "curl",
    "Postman Collection Link": "collection_link",
}


class ExcelAdapterError(ValueError):
    """The workbook could not be read as an API documentation template."""


def _require_openpyxl():
    try:
        import openpyxl  # noqa: PLC0415 - optional dependency, imported lazily
    except ImportError as error:  # pragma: no cover - environment-dependent
        raise ExcelAdapterError(
            "Reading an uploaded Excel definition requires openpyxl "
            "(listed in dev-requirements.txt). Install it with "
            "`pip install -r dev-requirements.txt`."
        ) from error
    return openpyxl


def _cell_text(value: Any) -> str:
    """Render a cell as trimmed text. Dates and numbers become their str form."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _sheet_rows(workbook, sheet_name: str) -> list[dict[str, str]]:
    """Read one sheet into header-keyed dicts, skipping fully blank rows."""
    if sheet_name not in workbook.sheetnames:
        raise ExcelAdapterError(
            f"workbook has no {sheet_name!r} sheet (found: {', '.join(workbook.sheetnames)})"
        )

    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = [_cell_text(cell) for cell in header_row]
    parsed: list[dict[str, str]] = []
    for raw_row in rows:
        values = [_cell_text(cell) for cell in raw_row]
        if not any(values):
            continue
        parsed.append(
            {
                header: values[index] if index < len(values) else ""
                for index, header in enumerate(headers)
                if header
            }
        )
    return parsed


def parse_workbook(path: str | Path) -> tuple[tuple[ApiDefinition, ...], tuple[str, ...]]:
    """Parse a template workbook into definitions plus non-fatal warnings.

    Warnings are structured and name the API ID. A lopsided sheet — more
    ``Error Request Body`` rows than ``Error Response`` rows, say — warns and
    pairs what it can; it never raises, because one malformed API must not take
    the batch down.
    """
    workbook_path = Path(path)
    if not workbook_path.exists():
        raise ExcelAdapterError(f"template workbook not found: {workbook_path}")

    openpyxl = _require_openpyxl()
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as error:
        raise ExcelAdapterError(
            f"{workbook_path.name} could not be opened as a workbook "
            f"({type(error).__name__})"
        ) from error

    try:
        overview = _sheet_rows(workbook, OVERVIEW_SHEET)
        payload_rows = _sheet_rows(workbook, PAYLOADS_SHEET)
        rule_rows = _sheet_rows(workbook, RULES_SHEET)
    finally:
        workbook.close()

    missing = [c for c in _OVERVIEW_COLUMNS if overview and c not in overview[0]]
    if missing:
        raise ExcelAdapterError(
            f"{OVERVIEW_SHEET} is missing required column(s): {', '.join(missing)}. "
            "Expected the 15-column template at api-docs/API_Documentation_Template.xlsx."
        )

    warnings: list[str] = []

    payloads_by_api: dict[str, list[SamplePayload]] = {}
    for row in payload_rows:
        api_id = row.get("API ID", "")
        if not api_id:
            continue
        payloads_by_api.setdefault(api_id, []).append(
            SamplePayload(
                payload_type=row.get("Payload Type", ""),
                # `na` (and n/a, -, empty) mean null. Never int() this unguarded:
                # `Request Body` and `Error Request Body` rows always carry it.
                response_status=_coerce_status(row.get("Response status")),
                sample_json=_parse_json_or_none(row.get("Sample JSON")),
            )
        )

    rules_by_api: dict[str, list[ApiRule]] = {}
    for row in rule_rows:
        api_id = row.get("API ID", "")
        if not api_id:
            continue
        # Free prose, stored verbatim and never parsed for metadata.
        rules_by_api.setdefault(api_id, []).append(
            ApiRule(
                category=row.get("Category", ""),
                description=row.get("Description", ""),
            )
        )

    definitions: list[ApiDefinition] = []
    for row in overview:
        values = {
            field: row.get(column, "")
            for column, field in _OVERVIEW_COLUMNS.items()
        }
        if not values["method"] or not values["path"]:
            warnings.append(
                f"excel-row-skipped api_id={values['api_id'] or '<blank>'!r}: "
                "HTTP Method and Endpoint Path are both required"
            )
            continue

        api_id = values["api_id"]
        payloads = tuple(payloads_by_api.get(api_id, ()))
        definitions.append(
            ApiDefinition(
                **values,
                payloads=payloads,
                rules=tuple(rules_by_api.get(api_id, ())),
            )
        )

        # GUARDRAIL — `Error Request Body` rows fire against UAT on every run.
        # They must provoke *validation or auth* errors, never state changes: a
        # DELETE with a malformed ID is a valid trigger; a DELETE with a real ID
        # that actually deletes something is not. Pairing is Nth-to-Nth in sheet
        # order; a mismatch warns and pairs the overlap rather than raising.
        error_requests = sum(
            1 for p in payloads if p.kind == PayloadType.normalize(PayloadType.ERROR_REQUEST_BODY)
        )
        error_responses = sum(
            1 for p in payloads if p.kind == PayloadType.normalize(PayloadType.ERROR_RESPONSE)
        )
        if error_requests != error_responses:
            warnings.append(
                f"error-sample-pairing-mismatch api_id={api_id!r} "
                f"error_request_bodies={error_requests} error_responses={error_responses}; "
                "paired the overlapping rows and ignored the remainder"
            )

    return tuple(definitions), tuple(warnings)


def load_excel_definitions(
    path: str | Path | None = None,
    *,
    api_ids: tuple[str, ...] | None = None,
) -> tuple[tuple[ApiDefinition, ...], tuple[str, ...]]:
    """Load definitions from the canonical template, optionally filtered by API ID."""
    definitions, warnings = parse_workbook(path or TEMPLATE_PATH)
    if api_ids is None:
        return definitions, warnings

    wanted = {str(a).strip() for a in api_ids}
    return tuple(d for d in definitions if d.api_id in wanted), warnings


def definition_with_base_url(
    definition: ApiDefinition,
    base_url: str,
) -> ApiDefinition:
    """Return a copy of ``definition`` whose base URL is ``base_url``."""
    return replace(definition, base_url=base_url)
